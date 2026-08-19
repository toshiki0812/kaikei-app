"""プランごとの試算結果を、数値だけのExcelブックに書き出す。

計算式は入れず、アプリのエンジンが出した数字をそのまま書く。
中身を書き換えても再計算はされないが、その代わり数字は必ずアプリと一致する。

使い方:  ./venv/bin/python scripts/export_values_workbook.py [出力先.xlsx] [月数]
"""
from __future__ import annotations

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as CL

import db
import simulation

F = "Arial"
HEAD = Font(name=F, bold=True, size=10)
BODY = Font(name=F, size=10)
TITLE = Font(name=F, bold=True, size=13)
NOTE = Font(name=F, size=9, color="666666")
ACCENT = Font(name=F, bold=True, size=10, color="1F5C3A")
HFILL = PatternFill("solid", fgColor="E8ECF0")
TFILL = PatternFill("solid", fgColor="F5F7F9")
THIN = Side(style="thin", color="D5DAE0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YEN = '#,##0;[Red]-#,##0;-'

# 世帯合計だけ income_total / credit_card_total という別名の列を持つ。
# 人別は income / credit_card なので、キーを分けておく。
SUMMARY = [("income_total", "収入合計"), ("total_expense", "支出合計"),
           ("net_cash_flow", "月次収支"), ("cash_balance", "現金残高"),
           ("investment_balance", "投資残高"), ("real_estate_value", "不動産評価額"),
           ("total_assets", "資産合計")]
PERSON_SUMMARY = [("income", lab) if key == "income_total" else (key, lab)
                  for key, lab in SUMMARY]
DETAIL = [("rent", "家賃"), ("credit_card_total", "クレカ"),
          ("investment_contribution", "投資拠出"), ("cash_sweep", "投資へ自動振替"),
          ("cash_shortfall_withdrawal", "投資から取り崩し"),
          ("other_expense", "その他既知支出"), ("other_cash_expense", "その他（現金）"),
          ("planned_income", "臨時収入"), ("planned_expense", "臨時支出"),
          ("real_estate_payment", "住宅ローン返済")]


def head(ws, row, labels, widths):
    for i, (label, w) in enumerate(zip(labels, widths), start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = HEAD
        c.fill = HFILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[CL(i)].width = w


def put(ws, row, col, value, fmt=YEN, font=BODY, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.border = BOX
    c.number_format = fmt
    if fill:
        c.fill = fill
    return c


def build_compare(wb, plans, frames, people, months):
    ws = wb.create_sheet("比較")
    ws["A1"] = f"プラン比較　{months[0]} 〜 {months[-1]}（{len(months)}ヶ月）"
    ws["A1"].font = TITLE
    labels = ["プラン", "資産合計", "現金残高", "投資残高", "不動産評価額", "収支累計"]
    labels += [f"{p['name']}の資産" for p in people]
    head(ws, 3, labels, [16, 15, 14, 14, 15, 14] + [14] * len(people))
    for i, plan in enumerate(plans):
        r = 4 + i
        df = frames[plan["id"]]
        last = df.iloc[-1]
        put(ws, r, 1, plan["name"], "@", HEAD)
        put(ws, r, 2, int(last["total_assets"]), font=ACCENT)
        for j, col in enumerate(["cash_balance", "investment_balance", "real_estate_value"], start=3):
            put(ws, r, j, int(last[col]))
        put(ws, r, 6, int(df["net_cash_flow"].sum()))
        for j, p in enumerate(people, start=7):
            put(ws, r, j, int(last[f"total_assets_p{p['id']}"]))
    n = len(plans)
    ws.cell(row=5 + n, column=1,
            value="各プランの詳細は、プラン名のシートを見てください。").font = NOTE
    ws.freeze_panes = "B4"
    return ws, 4 + n - 1


def build_plan_sheet(wb, plan, df, people, months):
    ws = wb.create_sheet(plan["name"][:31])
    ws["A1"] = f"{plan['name']}　月次の試算"
    ws["A1"].font = TITLE
    ws["A2"] = (plan["description"] or "")
    ws["A2"].font = NOTE

    labels = ["年月"] + [lab for _, lab in SUMMARY]
    for p in people:
        labels += [f"{p['name']}の{lab}" for _, lab in PERSON_SUMMARY]
    labels += [lab for _, lab in DETAIL]
    head(ws, 4, labels, [11] + [14] * (len(labels) - 1))

    for i, m in enumerate(months):
        r = 5 + i
        row = df.iloc[i]
        put(ws, r, 1, simulation.month_label(m), "@", BODY, TFILL)
        col = 2
        for key, _ in SUMMARY:
            put(ws, r, col, int(row[key]), fill=TFILL if key == "total_assets" else None)
            col += 1
        for p in people:
            for key, _ in PERSON_SUMMARY:
                put(ws, r, col, int(row[f"{key}_p{p['id']}"]))
                col += 1
        for key, _ in DETAIL:
            put(ws, r, col, int(row[key]))
            col += 1
    ws.freeze_panes = "B5"
    return ws


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "プラン比較.xlsx"
    n_months = int(sys.argv[2]) if len(sys.argv) > 2 else db.get_horizon_years() * 12

    plans = db.get_plans()
    people = db.get_people()
    frames = simulation.build_projections([p["id"] for p in plans], n_months)
    months = list(frames[plans[0]["id"]]["month"])

    wb = Workbook()
    wb.remove(wb.active)
    ws_cmp, last_row = build_compare(wb, plans, frames, people, months)
    for plan in plans:
        build_plan_sheet(wb, plan, frames[plan["id"]], people, months)

    # 資産合計の推移を1枚に重ねる
    ch = LineChart()
    ch.title = "資産合計の推移（世帯合計）"
    ch.y_axis.title = "円"
    ch.height, ch.width = 11, 26
    for plan in plans:
        s = wb[plan["name"][:31]]
        ch.add_data(Reference(s, min_col=8, min_row=4, max_row=4 + len(months)),
                    titles_from_data=True)
    ch.set_categories(Reference(wb[plans[0]["name"][:31]], min_col=1,
                                min_row=5, max_row=4 + len(months)))
    for i, ser in enumerate(ch.series):
        ser.tx.strRef.f = f"'{plans[i]['name'][:31]}'!$H$4"
        ser.graphicalProperties.line.width = 22000
        ser.smooth = False
    ws_cmp.add_chart(ch, f"A{last_row + 4}")

    wb.save(out)
    print(f"書き出しました: {out}")
    print(f"  {len(plans)}プラン × {len(months)}ヶ月／シート {len(wb.sheetnames)}枚")
    return frames, plans, months


if __name__ == "__main__":
    main()
