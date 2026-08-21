"""予実管理シートの計算式を、自前で評価してエンジンの出力と突き合わせる。

Excel も LibreOffice も無い環境で「式が正しいか」を確かめるための道具。
export_tracker.py が使う関数を IF/AND/OR/MAX/MIN/ROUND と四則演算に絞ってあるので、
それだけを解釈する小さな評価器で全セルを計算できる。
"""
from __future__ import annotations

import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook

CELL = re.compile(r"(?:(?P<sheet>[^\s!+\-*/(),<>=]+)!)?\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)")


def col_num(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


class Evaluator:
    """セルの式を必要になった順に評価していく（依存は再帰でたどる）。"""

    def __init__(self, path):
        self.wb = load_workbook(path)
        self.cache = {}
        self.stack = set()

    def raw(self, sheet, col, row):
        return self.wb[sheet].cell(row=row, column=col_num(col)).value

    def value(self, sheet, col, row):
        key = (sheet, col, row)
        if key in self.cache:
            return self.cache[key]
        if key in self.stack:
            raise RuntimeError(f"循環参照: {sheet}!{col}{row}")
        self.stack.add(key)
        v = self.raw(sheet, col, row)
        if isinstance(v, str) and v.startswith("="):
            v = self.eval_formula(v[1:], sheet)
        elif v is None:
            v = BLANK
        self.stack.discard(key)
        self.cache[key] = v
        return v

    def eval_formula(self, expr, cur_sheet):
        def sub(m):
            sheet = m.group("sheet") or cur_sheet
            val = self.value(sheet, m.group("col"), int(m.group("row")))
            return repr(val)
        # 式の中の "" は空欄を表す。0として計算でき、="" とも等しい値に置き換える
        # （自前の評価器は IF の両方の枝を計算してしまうため）
        expr = expr.replace('""', "_BLANK")
        py = CELL.sub(sub, expr)
        py = py.replace("<>", "!=").replace("^", "**")
        py = re.sub(r"(?<![<>!])=(?!=)", "==", py)
        py = py.replace("MAX(", "_max(").replace("MIN(", "_min(")
        py = py.replace("ROUND(", "_round(").replace("IF(", "_if(")
        py = py.replace("AND(", "_and(").replace("OR(", "_or(")
        return eval(py, EVAL_ENV)  # noqa: S307  式の語彙を絞ってあるため


class Blank(float):
    """空セル。Excelでは計算上0として扱われ、="" の比較では真になる。

    自前の評価器は IF の両方の枝を計算してしまうため、空欄が混ざると
    そのままでは型エラーになる。0として振る舞いつつ "" とも等しい値を用意する。
    """
    def __new__(cls):
        return super().__new__(cls, 0.0)

    def __eq__(self, other):
        return other == "" or other == 0
    def __ne__(self, other):
        return not self.__eq__(other)
    def __hash__(self):
        return hash("")
    def __repr__(self):
        return "_BLANK"


BLANK = Blank()


def _num(v):
    return 0 if v == "" or v is None else v


def _if(cond, a, b=False):
    return a if cond else b


def _round(v, digits=0):
    v = _num(v)
    # Excel の ROUND は 0.5 を常に切り上げ（Python の round は偶数丸め）
    import decimal
    return int(decimal.Decimal(str(v)).quantize(decimal.Decimal(1),
                                                rounding=decimal.ROUND_HALF_UP))


EVAL_ENV = {
    "_max": lambda *a: max(_num(x) for x in a),
    "_min": lambda *a: min(_num(x) for x in a),
    "_round": _round,
    "_if": _if,
    "_and": lambda *a: all(a),
    "_or": lambda *a: any(a),
    "_BLANK": BLANK,
    "__builtins__": {},
}


def main():
    path = sys.argv[1]
    n = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    ev = Evaluator(path)
    from scripts.export_tracker import BREAK, C_H_SUM, C_S_BR, C_S_SUM, C_T_BR, C_T_SUM
    from openpyxl.utils import get_column_letter as CL

    got = {"T": [], "S": [], "H": []}
    detail = {"T": [], "S": []}
    for i in range(n):
        r = 4 + i
        got["T"].append(ev.value("予実管理", CL(C_T_SUM), r))
        got["S"].append(ev.value("予実管理", CL(C_S_SUM), r))
        got["H"].append(ev.value("予実管理", CL(C_H_SUM), r))
        for who, base in (("T", C_T_BR), ("S", C_S_BR)):
            detail[who].append({
                name: ev.value("予実管理", CL(base + BREAK.index(name)), r)
                for name in ("現金残高", "投資残高", "不動産評価額")})
    return got, detail


if __name__ == "__main__":
    got, detail = main()
    print("評価できた月数:", len(got["T"]))
    print("最終月  俊来 %s / 新季 %s / 世帯 %s" %
          (format(got["T"][-1], ","), format(got["S"][-1], ","), format(got["H"][-1], ",")))
