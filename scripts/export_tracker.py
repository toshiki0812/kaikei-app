"""予実管理シートを作る。前提を書き換えれば全部が再計算される作りにする。

月ごとのブロックを縦に積み、収入・費用・残高を項目別に
「予算／実績／差異」で並べる。飲み・外食と自炊は独立した行にしてあり、
毎月そこが枠に収まったかを見るのがこのシートの主目的。

計算式は IF / AND / OR / MAX / MIN / ROUND / SUM / SUMIFS / COUNTIFS と
四則演算・セル参照だけで書く。使う関数を絞ってあるのは、Excelを持っていない
環境でも自前の評価器で全セルを検算できるようにするため（verify_tracker.py）。

使い方:  ./venv/bin/python scripts/export_tracker.py [出力先.xlsx] [月数]
"""
from __future__ import annotations

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as CL

import simulation

F = "Arial"
INK = "FF1A1A1A"
BLUE = "FF0000FF"        # 書き換えてよいセル
GREEN = "FF008000"       # 別シート参照
THIN = Side(style="thin", color="FFD5DAE0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HFILL = PatternFill("solid", fgColor="FF2C3E50")
SUBFILL = PatternFill("solid", fgColor="FFE8ECF0")
INFILL = PatternFill("solid", fgColor="FFFFF9E0")   # 実績の入力欄
YEN = '#,##0;[Red]-#,##0;"-"'

# ── 前提シートの番地。式から名前で引けるようにここで一元管理する ──
A = {}   # {キー: "前提!$B$3"} の形


def setup_sheet(wb, months):
    ws = wb.create_sheet("前提")
    ws.column_dimensions["A"].width = 30
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 15
    r = [1]

    def head(text):
        ws.cell(row=r[0], column=1, value=text).font = Font(name=F, bold=True, size=12)
        r[0] += 2

    def val(label, key, value, note="", col=2):
        ws.cell(row=r[0], column=1, value=label).font = Font(name=F, size=10)
        c = ws.cell(row=r[0], column=col, value=value)
        c.font = Font(name=F, size=10, color=BLUE)
        c.number_format = YEN if isinstance(value, int) else "General"
        c.border = BOX
        if note:
            ws.cell(row=r[0], column=col + 2, value=note).font = Font(
                name=F, size=9, color="FF888888")
        A[key] = f"前提!${CL(col)}${r[0]}"
        r[0] += 1

    ws["A1"] = "前提条件"
    ws["A1"].font = Font(name=F, bold=True, size=15)
    ws["C1"] = "青い数字は書き換えてOK。書き換えると予実管理シートが全部再計算されます。"
    ws["C1"].font = Font(name=F, size=10, color="FF888888")
    r[0] = 3

    head("■ 期間・運用")
    val("シミュレーション開始月", "start", months[0], "YYYY-MM")
    val("想定年利（%）", "rate", 5.0, "投資の運用利回り")
    r[0] += 1

    head("■ 俊来")
    val("収入（通常）", "T_income", 500000)
    val("収入（低収入期）", "T_income_low", 260000, "2026-10〜2027-03")
    val("低収入期の最終月", "T_low_end", "2027-03")
    val("家賃（今の家・2026-10のみ）", "T_rent_old", 167000)
    val("家賃（新居・同棲後）", "T_rent_new", 80000)
    val("クレカ（食費以外）", "T_card", 50000)
    val("　うち分割払いの上乗せ", "T_split", 29803, "2026-12で終了")
    val("　分割払いの最終月", "T_split_end", "2026-12")
    val("飲み・外食", "T_food_out", 30000)
    val("自炊", "T_food_home", 30000)
    val("投資拠出（通常）", "T_contrib", 100000)
    val("投資拠出を止める期間の最終月", "T_pause_end", "2027-03", "2026-11から")
    val("現金の上限（超えた分は投資へ）", "T_cap", 1000000)
    val("開始時の現金", "T_cash0", 2000000)
    val("開始時の投資", "T_inv0", 0)
    r[0] += 1

    head("■ 新季")
    val("収入（通常）", "S_income", 250000)
    val("収入（産休・育休中）", "S_income_leave", 150000)
    val("産休・育休の開始月", "S_leave_start", "2030-04")
    val("産休・育休の最終月", "S_leave_end", "2031-05")
    val("家賃", "S_rent", 80000)
    val("クレカ", "S_card", 100000)
    val("投資拠出", "S_contrib", 10000)
    val("現金の上限", "S_cap", 2000000)
    val("開始時の現金", "S_cash0", 0)
    val("開始時の投資", "S_inv0", 0)
    r[0] += 1

    head("■ 二人で共通")
    val("同棲の開始月", "move", "2026-11", "この月から新居の家賃・光熱費")
    val("光熱費の節約（一人あたり）", "util", 2500, "同棲した月から")
    val("第一子の継続費（一人あたり）", "child_m", 7500, "おむつ・ミルクなど")
    val("　継続費の開始月", "child_start", "2030-06")
    val("　継続費の最終月", "child_end", "2031-05")
    r[0] += 1

    head("■ 住宅ローン（一人あたり）")
    val("開始月", "loan_start", "2031-01")
    val("月々の返済額", "loan_pay", 100000)
    val("返済期間（ヶ月）", "loan_term", 420, "このシートの60ヶ月を超えるため終了は考慮しない")
    val("年間の値動き（%）", "loan_rate_tmp", 0.5)
    r[0] += 1

    head("■ 自動計算（さわらない）")
    ws.cell(row=r[0], column=1, value="投資の月利").font = Font(name=F, size=10)
    c = ws.cell(row=r[0], column=2, value=f"=(1+{A['rate']}/100)^(1/12)-1")
    c.font = Font(name=F, size=10, color=GREEN); c.number_format = "0.00000000"
    A["_inv_rate"] = f"前提!$B${r[0]}"
    r[0] += 1
    ws.cell(row=r[0], column=1, value="不動産の月利").font = Font(name=F, size=10)
    c2 = ws.cell(row=r[0], column=2, value=f"=(1+{A['loan_rate_tmp']}/100)^(1/12)-1")
    c2.font = Font(name=F, size=10, color=GREEN); c2.number_format = "0.00000000"
    A["_re_rate"] = f"前提!$B${r[0]}"
    r[0] += 2

    return ws


# ── イベント（単発・毎年）。予実管理シートから月ごとに拾う ──
ONCE = [   # (月, 対象, ラベル, 金額, 収入か)
    ("2026-11", "T", "違約金", 167000, False),
    ("2026-11", "T", "引越し費用", 600000, False),
    ("2027-04", "T", "C-LinC 1〜3月分報酬", 1100000, True),
    ("2027-09", "T", "婚約", 600000, False),
    ("2028-08", "T", "結婚式（費用）", 2000000, False),
    ("2028-08", "S", "結婚式（費用）", 2000000, False),
    ("2028-09", "T", "結婚式（ご祝儀）", 1000000, True),
    ("2028-09", "S", "結婚式（ご祝儀）", 1000000, True),
    ("2028-10", "T", "新婚旅行", 500000, False),
    ("2028-10", "S", "新婚旅行", 500000, False),
    ("2030-06", "T", "第一子・一時費用", 150000, False),
    ("2030-06", "S", "第一子・一時費用", 150000, False),
]
YEARLY = [  # (月, 対象, ラベル, 金額, 収入か)
    (2, "T", "住民税", 80000, False), (5, "T", "住民税", 80000, False),
    (8, "T", "住民税", 80000, False), (11, "T", "住民税", 80000, False),
    (3, "T", "所得税", 500000, False),
    (7, "S", "賞与（夏季）", 300000, True), (12, "S", "賞与（冬季）", 300000, True),
]


def events_sheet(wb):
    ws = wb.create_sheet("イベント")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws["A1"] = "イベント（単発）"
    ws["A1"].font = Font(name=F, bold=True, size=13)
    for i, h in enumerate(["年月", "対象", "内容", "金額", "収入/支出"], start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name=F, bold=True, size=10, color="FFFFFFFF")
        c.fill = HFILL
        c.border = BOX
    for j, (m, who, label, amt, is_in) in enumerate(ONCE):
        r = 4 + j
        for i, v in enumerate([m, "俊来" if who == "T" else "新季", label, amt,
                               "収入" if is_in else "支出"], start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=F, size=10, color=BLUE if i == 4 else INK)
            c.border = BOX
            if i == 4:
                c.number_format = YEN

    top = 4 + len(ONCE) + 2
    ws.cell(row=top, column=1, value="イベント（毎年くり返し）").font = Font(
        name=F, bold=True, size=13)
    for i, h in enumerate(["何月", "対象", "内容", "金額", "収入/支出"], start=1):
        c = ws.cell(row=top + 2, column=i, value=h)
        c.font = Font(name=F, bold=True, size=10, color="FFFFFFFF")
        c.fill = HFILL
        c.border = BOX
    for j, (mo, who, label, amt, is_in) in enumerate(YEARLY):
        r = top + 3 + j
        for i, v in enumerate([mo, "俊来" if who == "T" else "新季", label, amt,
                               "収入" if is_in else "支出"], start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=F, size=10, color=BLUE if i == 4 else INK)
            c.border = BOX
            if i == 4:
                c.number_format = YEN
    return ws


def event_expr(month, who, want_income):
    """その月・その人のイベント合計を、前提を参照する式の文字列で返す。

    SUMIFS を使わず、該当するイベントのセル参照を + でつなぐ。
    金額セルを直接参照するので、イベントシートの金額を書き換えれば反映される。
    """
    parts = []
    for j, (m, w, _l, _a, is_in) in enumerate(ONCE):
        if m == month and w == who and is_in == want_income:
            parts.append(f"イベント!$D${4 + j}")
    top = 4 + len(ONCE) + 2
    mo = int(month[5:])
    for j, (ym, w, _l, _a, is_in) in enumerate(YEARLY):
        if ym == mo and w == who and is_in == want_income:
            parts.append(f"イベント!$D${top + 3 + j}")
    return "+".join(parts) if parts else "0"




# ── 1ヶ月ぶんの行構成 ──────────────────────────────
# (区分, 項目, 種別)  種別: in=実績を手入力 / eq=式 / copy=予算をそのまま初期値に / calc=計算用（隠す）
BLOCK = [
    ("収入", "収入", "in"),
    ("収入", "臨時収入", "in"),
    ("収入", "収入計", "eq"),
    ("費用", "家賃", "copy"),
    ("費用", "クレカ（食費以外）", "in"),
    ("費用", "飲み・外食", "in"),
    ("費用", "自炊", "in"),
    ("費用", "食費計", "eq"),
    ("費用", "投資拠出", "in"),
    ("費用", "臨時支出", "in"),
    ("費用", "住宅ローン返済", "copy"),
    ("費用", "支出計", "eq"),
    ("振替", "投資へ自動振替", "eq"),
    ("振替", "投資から取り崩し", "eq"),
    ("残高", "現金残高", "in"),
    ("残高", "投資残高", "in"),
    ("残高", "不動産評価額", "copy"),
    ("残高", "資産合計", "eq"),
    ("計算用", "投資拠出（予定）", "calc"),
    ("計算用", "現金（振替前）", "calc"),
    ("計算用", "拠出の見送り", "calc"),
]
IDX = {name: i for i, (_g, name, _k) in enumerate(BLOCK)}
BLOCK_H = len(BLOCK) + 1          # 末尾に区切りの空行を1つ
SCORE_ROWS = 13                   # スコアボードが占める行数
HEAD_ROW = SCORE_ROWS + 2
FIRST_ROW = HEAD_ROW + 1

C_MONTH, C_GROUP, C_ITEM = 1, 2, 3
C_T, C_S, C_H = 4, 7, 10          # 各3列（予算／実績／差異）

COST_ITEMS = {"家賃", "クレカ（食費以外）", "飲み・外食", "自炊", "食費計",
              "投資拠出", "臨時支出", "住宅ローン返済", "支出計"}
INPUT_ITEMS = {"収入", "臨時収入", "クレカ（食費以外）", "飲み・外食", "自炊",
               "投資拠出", "臨時支出", "現金残高", "投資残高"}

FILL_IN = PatternFill("solid", fgColor="FFFFF9E0")     # 手入力
FILL_EQ = PatternFill("solid", fgColor="FFF2F5F8")     # 式（小計）
FILL_KEY = PatternFill("solid", fgColor="FFE8F1FB")    # 飲み・外食など注目行
FILL_MON = PatternFill("solid", fgColor="FF2C3E50")
GAP_COST = '[Red]+#,##0;-#,##0;"±0"'                   # 費用：使いすぎが赤
GAP_GAIN = '#,##0;[Red]-#,##0;"±0"'                    # 収入・残高：不足が赤


def row_of(i, item):
    """月インデックス i の項目 item がある行番号。"""
    return FIRST_ROW + i * BLOCK_H + IDX[item]


def budget_formulas(i, month, who, mc):
    """その月・その人の予算列の式を {項目: 式} で返す。"""
    A_ = A
    R = lambda name: f"{CL(C_T if who == 'T' else C_S)}{row_of(i, name)}"
    PV = lambda name: f"{CL(C_T if who == 'T' else C_S)}{row_of(i - 1, name)}"
    first = (i == 0)
    key = "T" if who == "T" else "S"
    cap, cash0, inv0 = A_[f"{key}_cap"], A_[f"{key}_cash0"], A_[f"{key}_inv0"]

    if who == "T":
        income = f'=IF({mc}<={A_["T_low_end"]},{A_["T_income_low"]},{A_["T_income"]})'
        rent = (f'=IF({mc}<{A_["move"]},{A_["T_rent_old"]},'
                f'IF({R("住宅ローン返済")}>0,0,{A_["T_rent_new"]}))')
        card = (f'={A_["T_card"]}+IF({mc}<={A_["T_split_end"]},{A_["T_split"]},0)'
                f'-IF({mc}>={A_["move"]},{A_["util"]},0)')
        food_out, food_home = f'={A_["T_food_out"]}', (
            f'={A_["T_food_home"]}+IF(AND({mc}>={A_["child_start"]},'
            f'{mc}<={A_["child_end"]}),{A_["child_m"]},0)')
        planned = f'=IF(AND({mc}>={A_["move"]},{mc}<={A_["T_pause_end"]}),0,{A_["T_contrib"]})'
    else:
        income = (f'=IF(AND({mc}>={A_["S_leave_start"]},{mc}<={A_["S_leave_end"]}),'
                  f'{A_["S_income_leave"]},{A_["S_income"]})')
        rent = f'=IF({R("住宅ローン返済")}>0,0,{A_["S_rent"]})'
        card = f'={A_["S_card"]}-IF({mc}>={A_["move"]},{A_["util"]},0)'
        food_out = "=0"
        food_home = (f'=IF(AND({mc}>={A_["child_start"]},{mc}<={A_["child_end"]}),'
                     f'{A_["child_m"]},0)')
        planned = f'={A_["S_contrib"]}'

    return {
        "収入": income,
        "臨時収入": "=" + event_expr(month, who, True),
        "収入計": f'={R("収入")}+{R("臨時収入")}',
        "家賃": rent,
        "クレカ（食費以外）": card,
        "飲み・外食": food_out,
        "自炊": food_home,
        "食費計": f'={R("飲み・外食")}+{R("自炊")}',
        "投資拠出": f'={R("投資拠出（予定）")}-{R("拠出の見送り")}',
        "臨時支出": "=" + event_expr(month, who, False),
        "住宅ローン返済": f'=IF({mc}>={A_["loan_start"]},{A_["loan_pay"]},0)',
        "支出計": (f'={R("家賃")}+{R("クレカ（食費以外）")}+{R("食費計")}'
                f'+{R("投資拠出")}+{R("臨時支出")}+{R("住宅ローン返済")}'),
        "投資へ自動振替": f'=MAX(0,{R("現金（振替前）")}-{cap})',
        "投資から取り崩し": (
            f'=IF({R("現金（振替前）")}-{R("投資へ自動振替")}+{R("拠出の見送り")}<0,'
            f'MIN(-({R("現金（振替前）")}-{R("投資へ自動振替")}+{R("拠出の見送り")}),'
            f'{inv0 if first else PV("投資残高")}),0)'),
        "現金残高": (f'={R("現金（振替前）")}-{R("投資へ自動振替")}'
                 f'+{R("拠出の見送り")}+{R("投資から取り崩し")}'),
        "投資残高": (f'=ROUND(({inv0 if first else PV("投資残高")}+{R("投資拠出")}'
                 f'+{R("投資へ自動振替")}-{R("投資から取り崩し")})*(1+{A_["_inv_rate"]}),0)'),
        "不動産評価額": (f'=ROUND(({"0" if first else PV("不動産評価額")}'
                   f'+{R("住宅ローン返済")})*(1+{A_["_re_rate"]}),0)'),
        "資産合計": f'={R("現金残高")}+{R("投資残高")}+{R("不動産評価額")}',
        "投資拠出（予定）": planned,
        "現金（振替前）": (f'={cash0 if first else PV("現金残高")}+{R("収入計")}'
                   f'-{R("家賃")}-{R("クレカ（食費以外）")}-{R("食費計")}'
                   f'-{R("投資拠出（予定）")}-{R("臨時支出")}-{R("住宅ローン返済")}'),
        "拠出の見送り": (f'=IF({R("現金（振替前）")}-{R("投資へ自動振替")}<0,'
                   f'MIN(-({R("現金（振替前）")}-{R("投資へ自動振替")}),'
                   f'{R("投資拠出（予定）")}),0)'),
    }
